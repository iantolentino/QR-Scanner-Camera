import cv2
import re
import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime
import time
import json
import os
import winsound  # Import for sound effects
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from tkinter import PhotoImage 

class AttendanceScanner:
    def __init__(self):
        # Attendance and camera variables
        self.camera = None
        self.is_on = False
        self.attendance = {}
        self.save_path = r"C:\AttendanceFiles\data"
        self.current_date = datetime.now().date()
        self.last_scan_time = 0  # Throttle re-scanning
        
        os.makedirs(os.path.dirname(self.save_path), exist_ok=True)
        
        # Initialize QR code detector
        self.qr_detector = cv2.QRCodeDetector()

        # Set up the main Tkinter window with background color #E5D0AC
        self.root = tk.Tk()
        self.root.title("Attendance Tracker")
        self.root.geometry("800x600")
        self.root.configure(bg="#E5D0AC")
        self.center()  # Center the main window
        self.root.iconbitmap(r"C:\AttendanceFiles\icons\calendar.ico") 

        # Main frame for all controls and logs
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # The camera feed widget is created but not packed (hidden)
        self.video_panel = ttk.Label(self.root, background="black")

        # Right Frame: Logs & Controls
        self.right_frame = ttk.Frame(self.main_frame)
        self.right_frame.pack(fill="both", expand=True)
        self.count_label = ttk.Label(self.right_frame, text="Workers Count: 0", font=("Helvetica", 16), foreground="#6D2323")
        self.count_label.pack(pady=10)

        # Log Text Widget with Scrollbar
        self.log_text = tk.Text(self.right_frame, wrap="word", font=("Helvetica", 12), height=15,
                                bg="#E5D0AC", fg="#6D2323", bd=0)
        self.log_text.pack(fill="both", expand=True, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(self.right_frame, command=self.log_text.yview)
        self.log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Button Frame
        self.button_frame = ttk.Frame(self.right_frame)
        self.button_frame.pack(pady=10)
        self.start_button = ttk.Button(self.button_frame, text="Start Scanner", command=self.turn_on)
        self.start_button.grid(row=0, column=0, padx=5, pady=5)
        self.stop_button = ttk.Button(self.button_frame, text="Stop Scanner", command=self.turn_off)
        self.stop_button.grid(row=0, column=1, padx=5, pady=5)
        self.download_excel_button = ttk.Button(self.button_frame, text="Download Excel", command=self.download_excel)
        self.download_excel_button.grid(row=1, column=0, padx=5, pady=5)
        self.download_json_button = ttk.Button(self.button_frame, text="Download JSON", command=self.download_json)
        self.download_json_button.grid(row=1, column=1, padx=5, pady=5)

        # Configure ttk style for buttons: background #6D2323 with white text
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", background="#6D2323", foreground="white", font=("Helvetica", 12), padding=10)
        style.map("TButton",
                  background=[("active", "#6D2323")],
                  foreground=[("active", "white")])

        # Status Label for notifications
        self.status_label = ttk.Label(self.right_frame, text="", font=("Helvetica", 12), foreground="#6D2323")
        self.status_label.pack(pady=5)

        self.root.protocol("WM_DELETE_WINDOW", self.close_app)
        self.check_auto_save()

    def center(self):
        # Center the window on the screen
        self.root.update_idletasks()
        width = 800
        height = 600
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def turn_on(self):
        if not self.is_on:
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                self.log_message("Error: Could not open camera.")
                return

            self.is_on = True
            self.status_label.config(text="Camera is ON")
            self.log_message("Camera has been turned ON.")
            self.update_frame()
        else:
            self.log_message("Camera is already ON.")

    def update_frame(self):
        if self.is_on:
            ret, frame = self.camera.read()
            if ret:
                # Process QR code detection every 2 seconds
                current_time = time.time()
                if current_time - self.last_scan_time > 2:
                    data, points, _ = self.qr_detector.detectAndDecode(frame)
                    if data:
                        id_match = re.search(r"id:\s*(\d+)", data)
                        name_match = re.search(r"name:\s*([\w-]+)", data)
                        if id_match and name_match:
                            worker_id = id_match.group(1)
                            worker_name = name_match.group(1)
                            now = datetime.now()
                            date_str = now.strftime("%Y-%m-%d")
                            if worker_id not in self.attendance:
                                self.attendance[worker_id] = {"name": worker_name, "daily_record": {}}
                            if date_str not in self.attendance[worker_id]["daily_record"]:
                                self.attendance[worker_id]["daily_record"][date_str] = {
                                    "time_in": None, "time_out": None, "overtime": 0, "worktime": 0
                                }
                            record = self.attendance[worker_id]["daily_record"][date_str]
                            # Play a general beep sound for any scan
                            winsound.Beep(1000, 200)  # Frequency: 1000 Hz, Duration: 200 ms
                            
                            if record["time_in"] is None:
                                record["time_in"] = now.strftime("%H:%M:%S")
                                self.log_message(f"ID: {worker_id}\nName: {worker_name}\nDate: {date_str}\nTime In: {record['time_in']}")
                                self.show_notification("Scan complete (Time In)")
                                # Play a specific sound for Time In
                                winsound.Beep(1500, 300)  # Higher pitch for Time In
                            elif record["time_out"] is None:
                                record["time_out"] = now.strftime("%H:%M:%S")
                                if now.hour >= 17:
                                    overtime_start = datetime.strptime(f"{date_str} 17:00:00", "%Y-%m-%d %H:%M:%S")
                                    overtime_hours = (now - overtime_start).total_seconds() / 3600
                                    record["overtime"] = max(0, overtime_hours)
                                record["worktime"] = self.calculate_work_time(record["time_in"], record["time_out"])
                                self.log_message(f"ID: {worker_id}\nName: {worker_name}\nDate: {date_str}\nTime In: {record['time_in']}\nTime Out: {record['time_out']}\nOvertime: {record['overtime']:.2f} hrs\nWorktime: {record['worktime']:.2f} hrs")
                                self.show_notification("Scan complete (Time Out)")
                                # Play a specific sound for Time Out
                                winsound.Beep(800, 300)  # Lower pitch for Time Out
                            self.last_scan_time = current_time
                            self.update_ui()
            self.root.after(30, self.update_frame)

    def turn_off(self):
        if self.is_on:
            self.is_on = False
            if self.camera:
                self.camera.release()
            self.log_message("Camera has been turned OFF.")
            self.status_label.config(text="Camera is OFF")

    def log_message(self, message):
        separator = "-" * 40
        self.log_text.insert(tk.END, f"{message}\n{separator}\n")
        self.log_text.see(tk.END)

    def show_notification(self, message):
        # Notification window using palette colors
        notification = tk.Toplevel(self.root)
        notification.title("Notification")
        notification.geometry("300x100")
        # Center notification over the main window
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 150
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 50
        notification.geometry(f"+{x}+{y}")
        notification.configure(bg="#E5D0AC")
        label = ttk.Label(notification, text=message, font=("Helvetica", 12), foreground="#6D2323", background="#E5D0AC")
        label.pack(expand=True, padx=10, pady=10)
        notification.after(2000, notification.destroy)

    def update_ui(self):
        self.count_label.config(text=f"Workers Count: {len(self.attendance)}")
        self.log_text.delete("1.0", tk.END)
        today = datetime.now().date().strftime("%Y-%m-%d")
        for worker_id, record in self.attendance.items():
            name = record["name"]
            if today in record["daily_record"]:
                times = record["daily_record"][today]
                time_in = times["time_in"] if times["time_in"] else "None"
                time_out = times["time_out"] if times["time_out"] else "None"
                overtime = times["overtime"]
                worktime = times["worktime"]
                self.log_text.insert(tk.END,
                    f"ID: {worker_id}\nName: {name}\nDate: {today}\nTime In: {time_in}\nTime Out: {time_out}\nOvertime: {overtime:.2f} hrs\nWorktime: {worktime:.2f} hrs\n{'-'*40}\n")
        self.log_text.see(tk.END)
        self.save_to_json()

    def save_to_json(self):
        json_path = f"{self.save_path}_{self.current_date.strftime('%Y%m%d')}.json"
        with open(json_path, 'w') as json_file:
            json.dump(self.attendance, json_file, indent=4)

    def calculate_work_time(self, time_in, time_out):
        if time_in and time_out:
            time_in_dt = datetime.strptime(time_in, "%H:%M:%S")
            time_out_dt = datetime.strptime(time_out, "%H:%M:%S")
            lunch_start = datetime.strptime("12:00:00", "%H:%M:%S")
            lunch_end = datetime.strptime("12:59:59", "%H:%M:%S")
            if time_out_dt.time() > lunch_start.time():
                time_out_dt = min(time_out_dt.replace(hour=12, minute=0, second=0), time_out_dt)
            if time_in_dt.time() < lunch_end.time():
                time_in_dt = max(time_in_dt.replace(hour=12, minute=59, second=59), time_in_dt)
            total_time = time_out_dt - time_in_dt
            return max(0, total_time.total_seconds() / 3600)
        return 0

    def download_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Attendance"
        today = datetime.now().date().strftime("%Y-%m-%d")
        headers = ["ID", "Name", "Time In", "Time Out", "Overtime (hrs)", "Worktime (hrs)", "Date"]
        ws.append(headers)
        for cell in ws["1:1"]:
            cell.font = Font(bold=True)
        for worker_id, record in self.attendance.items():
            if today in record["daily_record"]:
                times = record["daily_record"][today]
                ws.append([worker_id, record['name'], times["time_in"] or "None", times["time_out"] or "None",
                           f"{times['overtime']:.2f}", f"{times['worktime']:.2f}", today])
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb.save(file_path)
            self.show_notification("Excel file saved successfully.")

    def download_json(self):
        self.save_to_json()
        file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if file_path:
            with open(file_path, 'w') as json_file:
                json.dump(self.attendance, json_file, indent=4)
            self.show_notification("JSON file saved successfully.")

    def auto_save_excel(self):
        self.download_excel()

    def check_auto_save(self):
        def save_if_time():
            current_time = datetime.now()
            if current_time.hour == 22 and current_time.minute == 1:
                self.auto_save_excel()
            self.root.after(60000, save_if_time)
        save_if_time()

    def close_app(self):
        if self.is_on:
            self.turn_off()
        self.save_attendance_to_excel()
        self.root.destroy()

    def save_attendance_to_excel(self):
        file_path = os.path.join(self.save_path, "data.xlsx")
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Data"
            headers = ["ID", "Name", "Date", "Time In", "Time Out", "Overtime (hrs)", "Worktime (hrs)"]
            ws.append(headers)
        today = datetime.now().date().strftime("%Y-%m-%d")
        ws.append([today])
        for worker_id, record in self.attendance.items():
            for date, times in record["daily_record"].items():
                time_in = times["time_in"] or "None"
                time_out = times["time_out"] or "None"
                if time_in != "None":
                    time_in = datetime.strptime(time_in, "%H:%M:%S").strftime("%I:%M %p")
                if time_out != "None":
                    time_out = datetime.strptime(time_out, "%H:%M:%S").strftime("%I:%M %p")
                ws.append([worker_id, record['name'], date, time_in, time_out, f"{times['overtime']:.2f}", f"{times['worktime']:.2f}"])
        retries = 5
        for attempt in range(retries):
            try:
                wb.save(file_path)
                self.show_notification("Attendance data saved successfully.")
                break
            except PermissionError:
                if attempt < retries - 1:
                    time.sleep(1)
                else:
                    self.show_notification("Failed to save attendance data. Please close the file if it's open.")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    scanner = AttendanceScanner()
    scanner.run()
